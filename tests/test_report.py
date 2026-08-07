"""The printable loading plan and the spreadsheet.

A generated document is hard to assert on without re-implementing the
generator, so these check the things a broken one actually gets wrong: the file
is the format it claims, the pick list has a row per box, and the numbers on the
page came from the plan rather than from nowhere.
"""

from __future__ import annotations

import pytest

pytest.importorskip("matplotlib", reason="install the viz extra")
pytest.importorskip("openpyxl", reason="install the viz extra")

from openpyxl import load_workbook  # noqa: E402

from app.report import (  # noqa: E402
    item_list_xlsx,
    item_list_xlsx_bytes,
    loading_plan_pdf,
    loading_plan_pdf_bytes,
    paginate,
    pick_list_rows,
)
from core.solver_ep import solve  # noqa: E402
from tools.gen_demo import generate  # noqa: E402


@pytest.fixture
def solved():
    job = generate(vehicle_code="CNT-40DV", mix="mixed", fill=0.8, stops=3, seed=77)
    return job, solve(job)


def test_pdf_is_a_pdf(tmp_path, solved):
    job, plan = solved
    out = loading_plan_pdf(job, plan, tmp_path / "plan.pdf")
    data = out.read_bytes()
    assert data.startswith(b"%PDF-")
    assert data.rstrip().endswith(b"%%EOF")
    assert len(data) > 10_000


def test_pick_list_has_a_row_per_placed_box_in_order(solved):
    job, plan = solved
    rows = pick_list_rows(job, plan)
    assert len(rows) == len(plan.placements)
    assert [row[0] for row in rows] == [str(p.seq) for p in plan.placements]
    # The load order works from the closed end outwards. Distance to the doors
    # is not monotonic -- it depends on box length too -- but the x coordinate
    # is, and that is the invariant the sequence is built on.
    xs = [int(row[5].split(",")[0]) for row in rows]
    assert xs == sorted(xs)


def test_pagination_never_produces_a_document_with_no_pages():
    assert paginate([], size=10) == [[]]
    assert len(paginate([["x"]] * 10, size=4)) == 3
    assert len(paginate([["x"]] * 8, size=4)) == 2
    with pytest.raises(ValueError):
        paginate([["x"]], size=0)


def test_pdf_can_be_written_to_a_stream(solved):
    job, plan = solved
    assert loading_plan_pdf_bytes(job, plan).startswith(b"%PDF-")


def test_spreadsheet_lists_every_placement(tmp_path, solved):
    job, plan = solved
    out = item_list_xlsx(job, plan, tmp_path / "plan.xlsx")
    book = load_workbook(out)

    assert book.sheetnames[:2] == ["Summary", "Pick list"]
    sheet = book["Pick list"]
    assert sheet.max_row == len(plan.placements) + 1  # + header

    orders = [sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)]
    assert orders == [p.seq for p in plan.placements]
    assert orders == sorted(orders), "the pick list has to be in loading order"


def test_spreadsheet_summary_matches_the_plan(tmp_path, solved):
    job, plan = solved
    book = load_workbook(item_list_xlsx(job, plan, tmp_path / "plan.xlsx"))
    facts = {
        row[0]: row[1]
        for row in book["Summary"].iter_rows(min_col=1, max_col=2, values_only=True)
    }
    assert facts["Job"] == plan.job_id
    assert facts["Boxes placed"] == plan.metrics.placed
    assert facts["Boxes left behind"] == plan.metrics.unplaced
    assert facts["Volume utilisation"] == pytest.approx(
        plan.metrics.volume_utilization, abs=1e-4
    )


def test_left_behind_sheet_appears_only_when_needed(tmp_path):
    crowded = generate(vehicle_code="CNT-20DV", mix="pallets", fill=2.0, seed=3)
    plan = solve(crowded)
    assert plan.unplaced
    book = load_workbook(item_list_xlsx(crowded, plan, tmp_path / "a.xlsx"))
    assert "Left behind" in book.sheetnames
    assert book["Left behind"].max_row == len(plan.unplaced) + 1

    roomy = generate(vehicle_code="TIR-1360", mix="cartons", fill=0.2, seed=4)
    easy = solve(roomy)
    assert not easy.unplaced
    assert "Left behind" not in load_workbook(
        item_list_xlsx(roomy, easy, tmp_path / "b.xlsx")
    ).sheetnames


def test_spreadsheet_can_be_written_to_a_stream(solved):
    job, plan = solved
    # xlsx is a zip; the magic number is the cheapest honest check.
    assert item_list_xlsx_bytes(job, plan).startswith(b"PK\x03\x04")


def test_reports_survive_a_plan_that_placed_nothing(tmp_path):
    from core import catalog
    from core.models import Dims, Item, ItemType, Job

    huge = ItemType(sku="HUGE", name="huge", dims=Dims(9000, 9000, 9000), weight_g=1000)
    job = Job(
        job_id="empty", vehicle=catalog.vehicle("CNT-20DV"),
        items=(Item(uid=0, type=huge),),
    )
    plan = solve(job)
    assert plan.placements == ()

    assert loading_plan_pdf(job, plan, tmp_path / "e.pdf").exists()
    book = load_workbook(item_list_xlsx(job, plan, tmp_path / "e.xlsx"))
    assert book["Pick list"].max_row == 1
