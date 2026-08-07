"""Printable loading plan and spreadsheet item list.

The PDF is produced with matplotlib rather than an HTML-to-PDF engine.
WeasyPrint and its relatives need GTK or Qt native libraries, which turns
"install and run" into an afternoon on Windows; matplotlib is already a
dependency of the schematic renderer and draws both the diagrams and the tables
here. The document is plain and dense on purpose -- it goes on a clipboard, not
in a brochure.

Two pages:

1. The load: header, metrics, the validator's verdict, and side and top views.
2. The pick list: every box in loading order, so a person can work down it.

Both extras are optional; ``core`` still has no third-party dependencies.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from core.models import Job, Plan
from core.validator import validate
from tools.view import INK_MUTED, INK_PRIMARY, INK_SECONDARY, SURFACE, stop_style

#: Rows per pick-list page. Enough to be worth printing, few enough to read.
ROWS_PER_PAGE = 34


def _stamp() -> str:
    return datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")


def _verdict(job: Job, plan: Plan) -> tuple[str, str]:
    """(text, colour). An unchecked plan is never reported as clean."""
    report = validate(job, plan)
    if not report.violations:
        return "all checks pass", "#0a7a0a"
    return report.summary(), "#d03b3b"


def _draw_breakdown(figure, job: Job, plan: Plan, top: float) -> None:
    """Per-stop counts and what was left behind, under the drawings.

    A loading plan that only shows the picture answers "where does it go" and
    not "did everything come", which is the question the person at the doors
    actually has.
    """
    weights = {item.uid: item.weight_g for item in job.items}
    by_stop: dict[int, tuple[int, int]] = {}
    for p in plan.placements:
        boxes, weight = by_stop.get(p.stop, (0, 0))
        by_stop[p.stop] = (boxes + 1, weight + weights[p.item_uid])

    figure.text(0.06, top, "BY STOP", fontsize=7, color=INK_MUTED)
    y = top - 0.019
    for stop in sorted(by_stop):
        boxes, weight = by_stop[stop]
        colour, _ = stop_style(stop)
        figure.text(0.062, y, "■", fontsize=8, color=colour, va="center")
        figure.text(
            0.082, y,
            f"stop {stop}   {boxes} boxes   {weight / 1e6:.2f} t",
            fontsize=8, color=INK_SECONDARY, va="center",
        )
        y -= 0.017

    if plan.unplaced:
        reasons: dict[str, int] = {}
        for u in plan.unplaced:
            reasons[u.reason] = reasons.get(u.reason, 0) + 1
        detail = ", ".join(f"{count} {reason}" for reason, count in sorted(reasons.items()))
        figure.text(0.52, top, "LEFT BEHIND", fontsize=7, color=INK_MUTED)
        figure.text(
            0.522, top - 0.019,
            f"{len(plan.unplaced)} boxes did not fit ({detail})",
            fontsize=8, color="#d03b3b", va="center",
        )


def pick_list_rows(job: Job, plan: Plan) -> list[list[str]]:
    """One row per placed box, in loading order."""
    inner = job.vehicle.inner
    return [
        [
            str(p.seq), p.sku, str(p.stop),
            f"{p.dims.l}×{p.dims.w}×{p.dims.h}",
            p.orientation.name,
            f"{p.pos.x}, {p.pos.y}, {p.pos.z}",
            f"{inner.l - p.pos.x - p.dims.l}",
        ]
        for p in plan.placements
    ]


def paginate(rows: list[list[str]], size: int = ROWS_PER_PAGE) -> list[list[list[str]]]:
    """Split the pick list into printable pages, never returning none.

    Split out so the pagination can be tested directly. Counting pages by
    scanning the finished PDF tests matplotlib's object compression, not this.
    """
    if size < 1:
        raise ValueError("page size must be at least 1")
    if not rows:
        return [[]]
    return [rows[start : start + size] for start in range(0, len(rows), size)]


def _prepare(out: Path | BinaryIO) -> Path | BinaryIO:
    """Both writers take a path or a buffer; only a path needs a directory."""
    if isinstance(out, Path):
        out.parent.mkdir(parents=True, exist_ok=True)
    return out


def loading_plan_pdf(job: Job, plan: Plan, out: Path | BinaryIO) -> Path | BinaryIO:
    """Write the whole document to a path or an open binary stream."""
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    matplotlib.rcParams["font.family"] = ["DejaVu Sans"]

    # Late import: tools.view owns the drawing, and duplicating it here would
    # give the printed plan and the screen plan two ways to disagree.
    from tools.view import _draw_cog, _draw_panel

    _prepare(out)
    metrics = plan.metrics
    verdict_text, verdict_colour = _verdict(job, plan)
    inner = job.vehicle.inner

    with PdfPages(out) as pdf:
        # --- page 1: the load ------------------------------------------
        figure = plt.figure(figsize=(8.27, 11.69), facecolor=SURFACE)  # A4 portrait
        figure.suptitle(
            f"Loading plan — {plan.job_id}",
            x=0.06, y=0.965, ha="left", fontsize=15, color=INK_PRIMARY,
        )
        figure.text(
            0.06, 0.941,
            f"{job.vehicle.name}   ·   {inner.l} × {inner.w} × {inner.h} mm"
            f"   ·   {plan.algorithm}",
            ha="left", fontsize=8, color=INK_SECONDARY,
        )
        figure.text(0.06, 0.925, _stamp(), ha="left", fontsize=7, color=INK_MUTED)
        # On its own line, right-aligned: on the subtitle line it collided with
        # the vehicle description on any load with more than one violation.
        figure.text(
            0.94, 0.924, verdict_text,
            ha="right", fontsize=8, color=verdict_colour, fontweight="bold",
        )

        if metrics is not None:
            cells = [
                ("volume", f"{metrics.volume_utilization:.1%}"),
                ("payload", f"{metrics.weight_utilization:.1%}"),
                ("placed", f"{metrics.placed} / {metrics.placed + metrics.unplaced}"),
                ("balance", f"{metrics.cog_longitudinal_mm:+d} mm"),
            ]
            for index, (label, value) in enumerate(cells):
                x = 0.06 + index * 0.235
                figure.text(x, 0.882, label.upper(), fontsize=7, color=INK_MUTED)
                figure.text(x, 0.858, value, fontsize=14, color=INK_PRIMARY)

        # Size the panels from the vehicle's own proportions. A fixed height
        # leaves a container floating in a third of a page of white space,
        # because the drawing keeps its aspect ratio and the axes box does not.
        panel_w = 0.88
        page_aspect = 8.27 / 11.69

        def panel_height(extent_mm: int) -> float:
            return panel_w * page_aspect * (extent_mm / inner.l) + 0.055

        side_h = panel_height(inner.h)
        top_h = panel_height(inner.w)
        side_y = 0.815 - side_h
        top_y = side_y - 0.035 - top_h

        side = figure.add_axes((0.06, side_y, panel_w, side_h))
        top = figure.add_axes((0.06, top_y, panel_w, top_h))
        _draw_panel(side, job, plan, "x", "z", "sku")
        _draw_panel(top, job, plan, "x", "y", "sku")
        _draw_cog(side, job, plan, "z")
        _draw_cog(top, job, plan, "y")
        side.set_title("side elevation", loc="left", fontsize=8, color=INK_SECONDARY)
        top.set_title("top view", loc="left", fontsize=8, color=INK_SECONDARY)

        stops = sorted({p.stop for p in plan.placements})
        if len(stops) > 1:
            from matplotlib.patches import Patch

            # Above the axes, not inside them: inside, it sat on top of the
            # freight it was meant to explain.
            side.legend(
                handles=[
                    Patch(facecolor=stop_style(s)[0], hatch=stop_style(s)[1] or None,
                          edgecolor=SURFACE, label=f"stop {s}")
                    for s in stops
                ],
                loc="lower right", bbox_to_anchor=(1.0, 1.01), frameon=False,
                fontsize=7, labelcolor=INK_SECONDARY, ncols=len(stops),
            )

        _draw_breakdown(figure, job, plan, top_y - 0.045)
        figure.text(
            0.06, 0.045,
            "Load in the order on the following pages: deepest first, then bottom up.\n"
            "The dashed rectangle marks where the centre of gravity has to stay.",
            ha="left", fontsize=7.5, color=INK_SECONDARY,
        )
        pdf.savefig(figure)
        plt.close(figure)

        # --- pick list pages -------------------------------------------
        rows = pick_list_rows(job, plan)
        headers = ["#", "sku", "stop", "size mm", "turn", "position mm", "to doors"]

        start = 0
        for page in paginate(rows):
            figure = plt.figure(figsize=(8.27, 11.69), facecolor=SURFACE)
            figure.suptitle(
                f"Pick list — {plan.job_id}",
                x=0.06, y=0.96, ha="left", fontsize=13, color=INK_PRIMARY,
            )
            figure.text(
                0.06, 0.937,
                f"boxes {start + 1}–{start + len(page)} of {len(rows)}"
                if page else "nothing was placed",
                ha="left", fontsize=8, color=INK_SECONDARY,
            )
            start += len(page)
            axes = figure.add_axes((0.06, 0.05, 0.88, 0.86))
            axes.axis("off")
            if page:
                table = axes.table(
                    cellText=page, colLabels=headers, loc="upper center",
                    cellLoc="left", colLoc="left",
                    colWidths=[0.06, 0.20, 0.07, 0.20, 0.09, 0.24, 0.11],
                )
                table.auto_set_font_size(False)
                table.set_fontsize(7.5)
                table.scale(1, 1.35)
                for (row, _), cell in table.get_celld().items():
                    cell.set_edgecolor("#e1e0d9")
                    cell.set_linewidth(0.6)
                    if row == 0:
                        cell.set_text_props(color=INK_MUTED, fontsize=7)
                        cell.set_facecolor(SURFACE)
                    elif row % 2 == 0:
                        cell.set_facecolor("#f6f5f1")
            else:
                axes.text(0, 1, "nothing was placed", fontsize=9, color=INK_MUTED)
            pdf.savefig(figure)
            plt.close(figure)

        info = pdf.infodict()
        info["Title"] = f"LoadZa loading plan {plan.plan_id}"
        info["Subject"] = f"{job.vehicle.name}, {len(plan.placements)} boxes"
        info["Creator"] = "LoadZa"

    return out


def item_list_xlsx(job: Job, plan: Plan, out: Path | BinaryIO) -> Path | BinaryIO:
    """Write the pick list as a spreadsheet, plus a summary sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _prepare(out)
    inner = job.vehicle.inner
    metrics = plan.metrics
    verdict_text, _ = _verdict(job, plan)

    book = Workbook()
    summary = book.active
    summary.title = "Summary"
    facts = [
        ("Job", plan.job_id),
        ("Plan", plan.plan_id),
        ("Vehicle", f"{job.vehicle.name} ({job.vehicle.code})"),
        ("Loading space mm", f"{inner.l} × {inner.w} × {inner.h}"),
        ("Payload limit t", round(job.vehicle.max_payload_g / 1e6, 2)),
        ("Algorithm", plan.algorithm),
        ("Generated", _stamp()),
        ("Checks", verdict_text),
    ]
    if metrics is not None:
        facts += [
            ("Volume utilisation", round(metrics.volume_utilization, 4)),
            ("Payload utilisation", round(metrics.weight_utilization, 4)),
            ("Boxes placed", metrics.placed),
            ("Boxes left behind", metrics.unplaced),
            ("Centre of gravity, lengthwise mm", metrics.cog_longitudinal_mm),
            ("Centre of gravity, sideways mm", metrics.cog_lateral_mm),
            ("Solve time ms", metrics.solve_ms),
        ]
    for row, (label, value) in enumerate(facts, start=1):
        summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row, column=2, value=value)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 46

    sheet = book.create_sheet("Pick list")
    headers = [
        "Order", "SKU", "Stop", "Length mm", "Width mm", "Height mm",
        "X mm", "Y mm", "Z mm", "Orientation", "To doors mm",
    ]
    header_fill = PatternFill("solid", fgColor="EFEEEA")
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
    for row, p in enumerate(plan.placements, start=2):
        values = [
            p.seq, p.sku, p.stop,
            p.dims.l, p.dims.w, p.dims.h,
            p.pos.x, p.pos.y, p.pos.z,
            p.orientation.name,
            inner.l - p.pos.x - p.dims.l,
        ]
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row, column=column, value=value)
    sheet.freeze_panes = "A2"
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 13

    if plan.unplaced:
        left = book.create_sheet("Left behind")
        for column, title in enumerate(["Item", "SKU", "Reason"], start=1):
            cell = left.cell(row=1, column=column, value=title)
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row, u in enumerate(plan.unplaced, start=2):
            left.cell(row=row, column=1, value=u.item_uid)
            left.cell(row=row, column=2, value=u.sku)
            left.cell(row=row, column=3, value=u.reason)
        for column in "ABC":
            left.column_dimensions[column].width = 16

    book.save(out)
    return out


def loading_plan_pdf_bytes(job: Job, plan: Plan) -> bytes:
    buffer = BytesIO()
    loading_plan_pdf(job, plan, buffer)
    return buffer.getvalue()


def item_list_xlsx_bytes(job: Job, plan: Plan) -> bytes:
    buffer = BytesIO()
    item_list_xlsx(job, plan, buffer)
    return buffer.getvalue()
